"""Process rotor-model validation data and generate manuscript figures.

The workbook contains two repeated tests.  In each test the advance ratio is
first increased and then decreased.  This script treats the rising and falling
branches of both repeats as four experimental runs, pairs them by nominal
operating point, and plots one four-run mean marker per condition.  The source
workbook is read-only.

Outputs
-------
Image/fig8_rotor_model_comparison.{pdf,png}
Image/fig9_rotor_model_diagnostics.{pdf,png}
data_process/model_validation_metrics.csv
data_process/model_validation_table.tex
data_process/model_validation_summary.json
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib.lines import Line2D
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "data" / "旋转测试数据汇总表_升力标定实验残差.xlsx"
DEFAULT_IMAGE_DIR = ROOT / "Image"
DEFAULT_OUTPUT_DIR = ROOT / "data_process"

RPM_LEVELS = (3000, 4000, 5000)
ANGLE_LEVELS = (0, 15, 30, 45, 60, 75, 90)

# Seven samples from viridis: perceptually uniform and color-vision friendly.
ANGLE_COLORS = {
    angle: color
    for angle, color in zip(
        ANGLE_LEVELS,
        mpl.colormaps["viridis"](np.linspace(0.05, 0.95, len(ANGLE_LEVELS))),
    )
}
RPM_COLORS = {3000: "#0077BB", 4000: "#EE7733", 5000: "#009988"}
RPM_MARKERS = {3000: "o", 4000: "s", 5000: "^"}


def configure_matplotlib() -> None:
    """Set compact, publication-oriented matplotlib defaults."""
    mpl.rcParams.update(
        {
            "font.family": "serif",
            "font.serif": ["Times New Roman", "Times", "DejaVu Serif"],
            "mathtext.fontset": "custom",
            "mathtext.rm": "Times New Roman",
            "mathtext.it": "Times New Roman:italic",
            "mathtext.bf": "Times New Roman:bold",
            "font.size": 8.0,
            "axes.titlesize": 9.0,
            "axes.labelsize": 9.0,
            "xtick.labelsize": 7.5,
            "ytick.labelsize": 7.5,
            "legend.fontsize": 7.5,
            "figure.dpi": 150,
            "savefig.dpi": 400,
            "savefig.bbox": "tight",
            "savefig.pad_inches": 0.03,
            "axes.spines.top": True,
            "axes.spines.right": True,
            "axes.linewidth": 0.7,
            "xtick.direction": "in",
            "ytick.direction": "in",
            "xtick.top": True,
            "ytick.right": True,
            "xtick.minor.visible": True,
            "ytick.minor.visible": True,
            "xtick.major.width": 0.7,
            "ytick.major.width": 0.7,
            "xtick.minor.width": 0.5,
            "ytick.minor.width": 0.5,
            "axes.unicode_minus": False,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )


def load_residual_details(workbook_path: Path) -> pd.DataFrame:
    """Read the cached values in the workbook's residual-detail sheet."""
    if not workbook_path.exists():
        raise FileNotFoundError(f"Workbook not found: {workbook_path}")

    workbook = load_workbook(workbook_path, data_only=True, read_only=True)
    if "残差明细" not in workbook.sheetnames:
        raise KeyError("The workbook does not contain the sheet '残差明细'.")

    sheet = workbook["残差明细"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=5, max_row=5))]
    records = list(sheet.iter_rows(min_row=6, values_only=True))
    workbook.close()

    data = pd.DataFrame(records, columns=headers)
    rename = {
        "rpm": "rpm",
        "角度(°)": "angle_deg",
        "采样序号": "sample_id",
        "J_1（公式）": "j_1",
        "J_2（公式）": "j_2",
        "J_均值": "j_mean",
        "C_T_prop_1（标定）": "ct_1",
        "C_T_prop_2（标定）": "ct_2",
        "C_T_prop_实验均值": "ct_exp",
        "C_T_prop理论插值": "ct_theory",
        "残差": "residual",
        "绝对残差": "abs_residual",
        "相对残差": "relative_residual",
        "状态": "status",
    }
    missing = [name for name in rename if name not in data.columns]
    if missing:
        raise KeyError(f"Missing expected columns in 残差明细: {missing}")

    data = data.rename(columns=rename)
    numeric_columns = [
        "rpm",
        "angle_deg",
        "sample_id",
        "j_1",
        "j_2",
        "j_mean",
        "ct_1",
        "ct_2",
        "ct_exp",
        "ct_theory",
        "residual",
        "abs_residual",
        "relative_residual",
    ]
    for column in numeric_columns:
        data[column] = pd.to_numeric(data[column], errors="coerce")
    return data


def agreement_metrics(frame: pd.DataFrame) -> dict[str, float | int]:
    """Calculate prediction-agreement statistics for one data subset."""
    measured = frame["ct_exp"].to_numpy(dtype=float)
    predicted = frame["ct_theory"].to_numpy(dtype=float)
    residual = measured - predicted

    slope, intercept = np.polyfit(predicted, measured, 1)
    pearson_r = float(np.corrcoef(predicted, measured)[0, 1])
    sse = float(np.sum(residual**2))
    sst = float(np.sum((measured - measured.mean()) ** 2))
    r_squared = 1.0 - sse / sst
    covariance = float(np.mean((predicted - predicted.mean()) * (measured - measured.mean())))
    ccc = 2.0 * covariance / (
        float(np.var(predicted))
        + float(np.var(measured))
        + float((predicted.mean() - measured.mean()) ** 2)
    )
    nonzero = np.abs(predicted) > np.finfo(float).eps

    return {
        "n": int(len(frame)),
        "bias": float(np.mean(residual)),
        "mae": float(np.mean(np.abs(residual))),
        "rmse": float(np.sqrt(np.mean(residual**2))),
        "mean_relative_residual": float(np.mean(residual[nonzero] / predicted[nonzero])),
        "mare": float(np.mean(np.abs(residual[nonzero] / predicted[nonzero]))),
        "max_abs_residual": float(np.max(np.abs(residual))),
        "r_squared": float(r_squared),
        "pearson_r": pearson_r,
        "slope": float(slope),
        "intercept": float(intercept),
        "ccc": float(ccc),
    }


def aggregate_four_branches(data: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    """Average two repeats on the rising and falling branches at each condition.

    The sequence peak is detected independently for every rpm/angle block.  A
    rising sample ``s`` is paired with falling sample ``2 * peak - s``.  The
    turn-around sample is shared by both branches, as it is the only measured
    value at the maximum advance ratio.
    """
    rows: list[dict[str, float | int]] = []
    counts = {
        "candidate_nominal_conditions": 0,
        "valid_four_branch_conditions": 0,
        "missing_four_branch_conditions": 0,
        "outside_theory_range_conditions": 0,
        "shared_turnaround_conditions": 0,
    }

    required = ["j_1", "j_2", "ct_1", "ct_2"]
    for (rpm, angle), group in data.groupby(["rpm", "angle_deg"], sort=True):
        group = group.sort_values("sample_id").copy()
        measurable = group.dropna(subset=["j_mean"])
        if measurable.empty:
            continue
        peak_row = measurable.loc[measurable["j_mean"].idxmax()]
        peak_id = int(peak_row["sample_id"])
        rising = group.loc[
            (group["sample_id"] >= 2) & (group["sample_id"] <= peak_id)
        ]

        by_sample = group.set_index("sample_id", drop=False)
        for _, up in rising.iterrows():
            counts["candidate_nominal_conditions"] += 1
            down_id = 2 * peak_id - int(up["sample_id"])
            if down_id not in by_sample.index:
                counts["missing_four_branch_conditions"] += 1
                continue
            down = by_sample.loc[down_id]
            if isinstance(down, pd.DataFrame):
                down = down.iloc[0]

            branch_values = pd.concat([up[required], down[required]])
            if branch_values.isna().any():
                counts["missing_four_branch_conditions"] += 1
                continue
            if pd.isna(up["ct_theory"]) or pd.isna(down["ct_theory"]):
                counts["outside_theory_range_conditions"] += 1
                continue

            j_values = np.array([up["j_1"], up["j_2"], down["j_1"], down["j_2"]], dtype=float)
            ct_values = np.array([up["ct_1"], up["ct_2"], down["ct_1"], down["ct_2"]], dtype=float)
            j_mean = float(np.mean(j_values))
            ct_exp = float(np.mean(ct_values))
            ct_theory = float(np.mean([up["ct_theory"], down["ct_theory"]]))
            residual = ct_exp - ct_theory
            shared_turnaround = int(up["sample_id"]) == peak_id
            counts["shared_turnaround_conditions"] += int(shared_turnaround)
            counts["valid_four_branch_conditions"] += 1

            rows.append(
                {
                    "rpm": int(rpm),
                    "angle_deg": int(angle),
                    "nominal_level": int(up["sample_id"]) - 2,
                    "rising_sample_id": int(up["sample_id"]),
                    "falling_sample_id": down_id,
                    "j_mean": j_mean,
                    "j_sd": float(np.std(j_values, ddof=1)),
                    "ct_exp": ct_exp,
                    "ct_sd": float(np.std(ct_values, ddof=1)),
                    "ct_theory": ct_theory,
                    "residual": residual,
                    "abs_residual": abs(residual),
                    "relative_residual": residual / ct_theory,
                    "shared_turnaround": shared_turnaround,
                }
            )

    return pd.DataFrame(rows), counts


def build_metrics(valid: pd.DataFrame) -> pd.DataFrame:
    """Return one row per rotational speed plus an overall row."""
    rows: list[dict[str, float | int | str]] = []
    for rpm in RPM_LEVELS:
        metrics = agreement_metrics(valid.loc[valid["rpm"] == rpm])
        rows.append({"group": f"{rpm} rpm", **metrics})
    rows.append({"group": "Overall", **agreement_metrics(valid)})
    return pd.DataFrame(rows)


def repeatability_metrics(valid: pd.DataFrame) -> dict[str, float | int]:
    """Summarize the within-condition spread of the four branch values."""
    spread = valid["ct_sd"].to_numpy(dtype=float)
    mean_ct = valid["ct_exp"].abs().to_numpy(dtype=float)
    return {
        "n_conditions": int(len(valid)),
        "mean_within_condition_sd": float(np.mean(spread)),
        "root_mean_square_within_condition_sd": float(np.sqrt(np.mean(spread**2))),
        "mean_relative_within_condition_sd": float(np.mean(spread / mean_ct)),
        "p95_within_condition_sd": float(np.quantile(spread, 0.95)),
    }


def _save_figure(fig: mpl.figure.Figure, base_path: Path) -> None:
    fig.savefig(base_path.with_suffix(".pdf"))
    fig.savefig(base_path.with_suffix(".png"), dpi=400)
    plt.close(fig)


def plot_comparison(valid: pd.DataFrame, image_dir: Path) -> None:
    """Generate theory/experiment comparison and residual panels by rpm."""
    fig, axes = plt.subplots(
        nrows=3,
        ncols=2,
        figsize=(7.15, 7.85),
        sharex=True,
        gridspec_kw={"width_ratios": [1.08, 1.0], "hspace": 0.18, "wspace": 0.22},
    )

    j_min, j_max = valid["j_mean"].min(), valid["j_mean"].max()
    ct_min = min(valid["ct_exp"].min(), valid["ct_theory"].min())
    ct_max = max(valid["ct_exp"].max(), valid["ct_theory"].max())
    ct_padding = 0.06 * (ct_max - ct_min)
    residual_limit = 1.08 * valid["residual"].abs().max()

    panel_index = 0
    for row_index, rpm in enumerate(RPM_LEVELS):
        comparison_ax, residual_ax = axes[row_index]
        rpm_data = valid.loc[valid["rpm"] == rpm]

        for angle in ANGLE_LEVELS:
            subset = rpm_data.loc[rpm_data["angle_deg"] == angle].sort_values("j_mean")
            if subset.empty:
                continue
            color = ANGLE_COLORS[angle]

            comparison_ax.plot(
                subset["j_mean"],
                subset["ct_theory"],
                color=color,
                linewidth=1.15,
                zorder=2,
            )
            comparison_ax.errorbar(
                subset["j_mean"],
                subset["ct_exp"],
                yerr=subset["ct_sd"],
                fmt="o",
                markersize=3.8,
                markerfacecolor="white",
                markeredgecolor=color,
                markeredgewidth=0.8,
                ecolor=color,
                elinewidth=0.55,
                capsize=1.5,
                capthick=0.55,
                zorder=3,
            )
            residual_ax.scatter(
                subset["j_mean"],
                subset["residual"],
                s=16,
                facecolors=color,
                edgecolors="white",
                linewidths=0.25,
                alpha=0.9,
                zorder=2,
            )

        comparison_ax.set_ylim(ct_min - ct_padding, ct_max + ct_padding)
        residual_ax.set_ylim(-residual_limit, residual_limit)
        residual_ax.axhline(0.0, color="0.25", linewidth=0.8, linestyle="--", zorder=1)

        for column_index, ax in enumerate((comparison_ax, residual_ax)):
            panel_label = chr(ord("a") + panel_index)
            panel_index += 1
            ax.text(
                0.015,
                0.96,
                f"({panel_label}) {rpm} rpm",
                transform=ax.transAxes,
                ha="left",
                va="top",
                fontsize=8.5,
                fontweight="bold",
            )
            ax.grid(axis="y", color="0.88", linewidth=0.45, linestyle=":", zorder=0)
            ax.set_xlim(j_min - 0.01, j_max + 0.01)

        comparison_ax.set_ylabel(r"$C_T$")
        residual_ax.set_ylabel(r"$C_{T,\mathrm{exp}}-C_{T,\mathrm{theory}}$")

    axes[-1, 0].set_xlabel(r"Advance ratio, $J$")
    axes[-1, 1].set_xlabel(r"Advance ratio, $J$")

    angle_handles = [
        Line2D([0], [0], color=ANGLE_COLORS[angle], lw=1.8, label=rf"${angle}^\circ$")
        for angle in ANGLE_LEVELS
    ]
    style_handles = [
        Line2D([0], [0], color="0.15", lw=1.25, label="Theory"),
        Line2D(
            [0],
            [0],
            marker="o",
            color="0.15",
            markerfacecolor="white",
            markeredgewidth=0.9,
            lw=0,
            label=r"Experiment mean $\pm$ SD",
        ),
    ]
    fig.legend(
        handles=angle_handles,
        title=r"Disk angle of attack, $\alpha_p$",
        loc="lower center",
        bbox_to_anchor=(0.36, -0.002),
        ncol=4,
        frameon=False,
        columnspacing=1.2,
        handlelength=1.7,
        title_fontsize=8,
    )
    fig.legend(
        handles=style_handles,
        loc="lower center",
        bbox_to_anchor=(0.80, 0.015),
        ncol=1,
        frameon=False,
        handlelength=1.8,
        labelspacing=0.35,
    )
    fig.subplots_adjust(left=0.095, right=0.99, top=0.995, bottom=0.115)
    _save_figure(fig, image_dir / "fig8_rotor_model_comparison")


def plot_diagnostics(valid: pd.DataFrame, metrics: pd.DataFrame, image_dir: Path) -> None:
    """Generate the parity plot and angle-dependent RMSE diagnostic."""
    fig, (parity_ax, rmse_ax) = plt.subplots(
        1, 2, figsize=(7.15, 3.15), gridspec_kw={"wspace": 0.28}
    )

    for rpm in RPM_LEVELS:
        subset = valid.loc[valid["rpm"] == rpm]
        parity_ax.scatter(
            subset["ct_theory"],
            subset["ct_exp"],
            s=15,
            marker=RPM_MARKERS[rpm],
            facecolors="none",
            edgecolors=RPM_COLORS[rpm],
            linewidths=0.65,
            alpha=0.72,
            label=f"{rpm} rpm",
        )

    lower = min(valid["ct_theory"].min(), valid["ct_exp"].min())
    upper = max(valid["ct_theory"].max(), valid["ct_exp"].max())
    padding = 0.05 * (upper - lower)
    line_limits = (lower - padding, upper + padding)
    parity_ax.plot(line_limits, line_limits, color="0.2", lw=1.0, linestyle="--", label="1:1 line")
    parity_ax.set_xlim(line_limits)
    parity_ax.set_ylim(line_limits)
    parity_ax.set_aspect("equal", adjustable="box")
    parity_ax.set_xlabel(r"Theoretical $C_T$")
    parity_ax.set_ylabel(r"Experimental $C_T$")
    parity_ax.grid(color="0.88", linewidth=0.45, linestyle=":")
    parity_ax.legend(frameon=False, loc="upper left", handletextpad=0.4)

    overall = metrics.loc[metrics["group"] == "Overall"].iloc[0]
    parity_ax.text(
        0.98,
        0.03,
        "\n".join(
            [
                rf"$C_{{T,\mathrm{{exp}}}}={overall['slope']:.3f}C_{{T,\mathrm{{theory}}}}{overall['intercept']:+.4f}$",
                rf"$R^2={overall['r_squared']:.3f}$, CCC$={overall['ccc']:.3f}$",
                rf"$N={int(overall['n'])}$",
            ]
        ),
        transform=parity_ax.transAxes,
        ha="right",
        va="bottom",
        fontsize=7.3,
        bbox={"facecolor": "white", "edgecolor": "0.75", "linewidth": 0.5, "alpha": 0.9},
    )

    for rpm in RPM_LEVELS:
        values = []
        for angle in ANGLE_LEVELS:
            subset = valid.loc[(valid["rpm"] == rpm) & (valid["angle_deg"] == angle)]
            values.append(float(np.sqrt(np.mean(subset["residual"].to_numpy() ** 2))))
        rmse_ax.plot(
            ANGLE_LEVELS,
            values,
            color=RPM_COLORS[rpm],
            marker=RPM_MARKERS[rpm],
            markersize=4.2,
            markerfacecolor="white",
            markeredgewidth=0.8,
            linewidth=1.25,
            label=f"{rpm} rpm",
        )

    rmse_ax.set_xlabel(r"Disk angle of attack, $\alpha_p$ ($^\circ$)")
    rmse_ax.set_ylabel("RMSE of $C_T$")
    rmse_ax.set_xticks(ANGLE_LEVELS)
    rmse_ax.set_ylim(bottom=0)
    rmse_ax.grid(axis="y", color="0.88", linewidth=0.45, linestyle=":")
    rmse_ax.legend(frameon=False, loc="upper right")

    parity_ax.text(0.01, 1.02, "(a)", transform=parity_ax.transAxes, fontweight="bold")
    rmse_ax.text(0.01, 1.02, "(b)", transform=rmse_ax.transAxes, fontweight="bold")
    fig.subplots_adjust(left=0.095, right=0.99, top=0.96, bottom=0.17)
    _save_figure(fig, image_dir / "fig9_rotor_model_diagnostics")


def write_outputs(
    data: pd.DataFrame,
    valid: pd.DataFrame,
    metrics: pd.DataFrame,
    aggregation_counts: dict[str, int],
    output_dir: Path,
) -> None:
    """Write auditable numeric summaries and the LaTeX table body."""
    metrics.to_csv(output_dir / "model_validation_metrics.csv", index=False, encoding="utf-8-sig")
    valid.to_csv(output_dir / "model_validation_four_branch_mean.csv", index=False, encoding="utf-8-sig")

    repeatability = repeatability_metrics(valid)
    status_counts = {str(key): int(value) for key, value in data["status"].value_counts().items()}
    residual = valid["residual"].to_numpy(dtype=float)
    overall = metrics.loc[metrics["group"] == "Overall"].iloc[0]
    summary = {
        "source_workbook": str(DEFAULT_WORKBOOK.relative_to(ROOT)),
        "status_counts": status_counts,
        "four_branch_aggregation": aggregation_counts,
        "aggregation_definition": (
            "Two repeated tests multiplied by rising/falling branches; one mean per nominal condition. "
            "The turn-around point is shared by the rising and falling branches."
        ),
        "negative_residual_fraction": float(np.mean(residual < 0)),
        "absolute_residual_within_0_005_fraction": float(np.mean(np.abs(residual) <= 0.005)),
        "absolute_residual_within_0_01_fraction": float(np.mean(np.abs(residual) <= 0.01)),
        "repeatability": repeatability,
        "overall": {key: (int(value) if key == "n" else float(value)) for key, value in overall.items() if key != "group"},
    }
    (output_dir / "model_validation_summary.json").write_text(
        json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
    )

    rows = []
    for _, row in metrics.iterrows():
        label = "全部" if row["group"] == "Overall" else row["group"].replace(" rpm", "")
        rows.append(
            f"{label} & {int(row['n'])} & {row['bias']:.5f} & {row['mae']:.5f} & "
            f"{row['rmse']:.5f} & {100 * row['mare']:.2f}\\% & {row['r_squared']:.3f} \\\\"
        )
    table_tex = "\n".join(
        [
            r"\begin{center}",
            r"\vbox{\centering{\small 表4\quad 旋翼升力模型精度统计", 
            r"\\ Table\,4 \quad Accuracy statistics of the rotor lift model}\vskip2mm",
            r"{\scriptsize\centering",
            r"\resizebox{\columnwidth}{!}{%",
            r"\begin{tabular}{crrrrrr}",
            r"\toprule",
            r"转速/(r/min) & $N$ & Bias & MAE & RMSE & MARE & $R^2$ \\",
            r"\midrule",
            *rows,
            r"\bottomrule",
            r"\end{tabular}}}",
            r"}",
            r"\end{center}",
        ]
    )
    (output_dir / "model_validation_table.tex").write_text(table_tex + "\n", encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK, help="Input xlsx workbook")
    parser.add_argument("--image-dir", type=Path, default=DEFAULT_IMAGE_DIR, help="Figure output directory")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Numeric/table output directory")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    image_dir = args.image_dir.resolve()
    output_dir = args.output_dir.resolve()
    image_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)

    configure_matplotlib()
    data = load_residual_details(workbook_path)
    valid, aggregation_counts = aggregate_four_branches(data)
    if valid.empty:
        raise ValueError("No valid four-branch operating conditions were produced.")

    metrics = build_metrics(valid)
    plot_comparison(valid, image_dir)
    plot_diagnostics(valid, metrics, image_dir)
    write_outputs(data, valid, metrics, aggregation_counts, output_dir)

    print(metrics[["group", "n", "bias", "mae", "rmse", "mare", "r_squared"]].to_string(index=False))
    print(json.dumps(aggregation_counts, indent=2))
    print(f"Figures written to: {image_dir}")
    print(f"Tables and summaries written to: {output_dir}")


if __name__ == "__main__":
    main()
